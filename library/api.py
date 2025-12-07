from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
import pyotp
from rest_framework import status
from django.utils import timezone
from datetime import timedelta
import random
from rest_framework.viewsets import GenericViewSet
from rest_framework import mixins, viewsets
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.contrib.auth import authenticate, login, logout
from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import serializers
from django.db.models import Avg, Count, Max, Min
from library.models import Author, Book, Genre, Reader, BookInstance
from library.serializers import AuthorSerializer, BookSerializer, GenreSerializer, ReaderSerializer, BookInstanceSerializer

class AuthorViewSet(mixins.ListModelMixin, 
                   mixins.CreateModelMixin, 
                   mixins.RetrieveModelMixin,
                   mixins.UpdateModelMixin,
                   mixins.DestroyModelMixin,
                   GenericViewSet):
    queryset = Author.objects.all()
    serializer_class = AuthorSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.AllowAny()]
    
    class StatsSerializer(serializers.Serializer):
        count = serializers.IntegerField()
        avg_books = serializers.FloatField()
        max_books = serializers.IntegerField()
        min_books = serializers.IntegerField()
        total_books = serializers.IntegerField()
        
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_stats(self, request, *args, **kwargs):
 
        total_authors = Author.objects.count()
        
        author_stats = Author.objects.annotate(
            book_count=Count('book') 
        ).aggregate(
            count=Count("*"),
            avg_books=Avg("book_count"),
            max_books=Max("book_count"),
            min_books=Min("book_count"),
        )
        
        total_books = Book.objects.count()
     
        stats = {
            'count': total_authors,
            'avg_books': author_stats['avg_books'],
            'max_books': author_stats['max_books'],
            'min_books': author_stats['min_books'],
            'total_books': total_books
        }
        
        serializer = self.StatsSerializer(instance=stats)
        return Response(serializer.data)

   

class GenreViewSet(mixins.ListModelMixin, 
                  mixins.CreateModelMixin, 
                  mixins.RetrieveModelMixin,
                  mixins.UpdateModelMixin,
                  mixins.DestroyModelMixin,
                  GenericViewSet):
    queryset = Genre.objects.all()
    serializer_class = GenreSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.AllowAny()]
    
    class GenreStatsSerializer(serializers.Serializer):
        total_genres = serializers.IntegerField()
        total_books = serializers.IntegerField()
        
        most_popular_genre = serializers.CharField()
        books_in_popular_genre = serializers.IntegerField()
        
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_genre_stats(self, request, *args, **kwargs):
        genre_stats = Genre.objects.annotate(
            book_count=Count('book')  
        ).order_by('-book_count')
        
        total_genres = genre_stats.count()
        total_books = Book.objects.count()
        
        
        most_popular = genre_stats.first()
        
        stats = {
            'total_genres': total_genres,
            'total_books': total_books,
            'most_popular_genre': most_popular.name,
            'books_in_popular_genre': most_popular.book_count
        }
        
        serializer = self.GenreStatsSerializer(instance=stats)
        return Response(serializer.data)
    

class BookViewSet(mixins.ListModelMixin, 
                 mixins.CreateModelMixin, 
                 mixins.RetrieveModelMixin,
                 mixins.UpdateModelMixin,
                 mixins.DestroyModelMixin,
                 GenericViewSet):
    queryset = Book.objects.all()
    serializer_class = BookSerializer
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.AllowAny()]
    class BookStatsSerializer(serializers.Serializer):
        total_books = serializers.IntegerField()
        avg_publication_year = serializers.FloatField()
        oldest_book_year = serializers.IntegerField()
        newest_book_year = serializers.IntegerField()
        most_popular_author = serializers.CharField()
        books_by_popular_author = serializers.IntegerField()
  
    @action(detail=False, methods=['GET'], url_path='export-excel')
    def export_books_to_excel(self, request, *args, **kwargs):
        books = Book.objects.all()
       
        wb = Workbook()
        ws = wb.active
        ws.title = "Книги библиотеки"
        
        headers = ['Название', 'Автор', 'Жанр', 'Год издания', 'Описание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, book in enumerate(books, 2):
            ws.cell(row=row, column=1, value=book.name)
            ws.cell(row=row, column=2, value=book.author.nameAuthor)
            ws.cell(row=row, column=3, value=book.genres.name )
            ws.cell(row=row, column=4, value=book.publication_year)
            ws.cell(row=row, column=5, value=book.description)
        
        # размерность столбцов 
        for column in ws.columns: 
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="library_books.xlsx"'
        
        return response
     
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_book_stats(self, request, *args, **kwargs):
        book_stats = Book.objects.aggregate(
            total_books=Count("*"),
            avg_publication_year=Avg("publication_year"),
            oldest_book_year=Min("publication_year"),
            newest_book_year=Max("publication_year"),
        )
        author_stats = Author.objects.annotate(
            book_count=Count('book')
        ).order_by('-book_count').first() 
        stats = {
            'total_books': book_stats['total_books'],
            'avg_publication_year': round(book_stats['avg_publication_year'], 1),
            'oldest_book_year': book_stats['oldest_book_year'],
            'newest_book_year': book_stats['newest_book_year'],
            'most_popular_author': author_stats.nameAuthor,
            'books_by_popular_author': author_stats.book_count,
        }
        serializer = self.BookStatsSerializer(instance=stats)
        return Response(serializer.data)
  

class UserViewSet(viewsets.GenericViewSet):
    permission_classes = [permissions.AllowAny]

    @action(detail=False, url_path="info", methods=["GET"])
    def get_info(self, *args, **kwargs):
        data = {
            'username': self.request.user.username,
            'is_authenticated': self.request.user.is_authenticated,
            'is_staff': self.request.user.is_staff,
        }

        if self.request.user.is_authenticated:
            data.update({
                'second': self.request.session.get('second') or False, 
            })

        return Response(data)

    @action(detail=False, url_path="login", methods=["POST"])
    def login_user(self, request, *args, **kwargs):
        username = self.request.data['username']
        password = self.request.data['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return Response({
                'success': True,
            })

        return Response({
            'success': False,
        })
    
    @action(url_path="get-totp", methods=['GET'], detail=False)
    def get_totp(self, *args, **kwargs):
        reader = Reader.objects.get(user=self.request.user)
        reader.totp_key = pyotp.random_base32()
        reader.save()
        url = pyotp.totp.TOTP(reader.totp_key).provisioning_uri(
            name=self.request.user.username, issuer_name="library"
        )

        return Response({
            "url": url
        })
    

    @action(detail=False, url_path="second-login", methods=["POST"])
    def second_login(self, *args, **kwargs):
        reader = Reader.objects.get(user=self.request.user) 
        key = reader.totp_key  
        t = pyotp.totp.TOTP(key)
        input_code = self.request.data.get('key', '')
        if input_code == t.now():
            self.request.session['second'] = True
            return Response({
                'success': True,
            })
    
    @action(url_path="logout", methods=['POST'], detail=False)
    def logout_user(self, *args, **kwargs):
        logout(self.request)

        return Response({
            "status": "success"
        })
    

class ReaderViewSet(mixins.ListModelMixin, 
                   mixins.CreateModelMixin, 
                   mixins.RetrieveModelMixin,
                   mixins.UpdateModelMixin,
                   mixins.DestroyModelMixin,
                   GenericViewSet):
    queryset = Reader.objects.all()
    serializer_class = ReaderSerializer
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    def get_queryset(self):
        qs = super().get_queryset()
        
        if self.request.user.is_superuser:
            return qs
        
        if self.request.user.is_authenticated:
            return qs.filter(user=self.request.user)
        
        return Reader.objects.none()

class BookInstanceViewSet(mixins.ListModelMixin, 
                         mixins.CreateModelMixin, 
                         mixins.RetrieveModelMixin,
                         mixins.UpdateModelMixin,
                         mixins.DestroyModelMixin,
                         GenericViewSet):
    queryset = BookInstance.objects.all()
    serializer_class = BookInstanceSerializer
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]

    
    def get_queryset(self):
        qs = super().get_queryset()
        
        if self.request.user.is_superuser:
            return qs
        
        if self.request.user.is_authenticated:
        
            try:
                reader = Reader.objects.get(user=self.request.user)
                return qs.filter(borrower=reader)
            except Reader.DoesNotExist:
                return BookInstance.objects.none()
        
        return BookInstance.objects.none()
    
    def perform_create(self, serializer):
       
        if self.request.user.is_authenticated:
            try:
                reader = Reader.objects.get(user=self.request.user)
                serializer.save(borrower=reader)
            except Reader.DoesNotExist:
                serializer.save()